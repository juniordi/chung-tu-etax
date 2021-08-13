from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import os, time, base64, datetime
# import Action chains 
from selenium.webdriver.common.action_chains import ActionChains
import os
import openpyxl
from xml.etree import ElementTree
import urllib.request, json, sys
#file goc: file-chrome.py
def get_downloaded_files(driver):

  if not driver.current_url.startswith("chrome://downloads"):
    driver.get("chrome://downloads/")

  return driver.execute_script( \
    "return downloads.Manager.get().items_   "
    "  .filter(e => e.state === 'COMPLETE')  "
    "  .map(e => e.filePath || e.file_path); " )


def get_file_content(driver, path):

  elem = driver.execute_script( \
    "var input = window.document.createElement('INPUT'); "
    "input.setAttribute('type', 'file'); "
    "input.hidden = true; "
    "input.onchange = function (e) { e.stopPropagation() }; "
    "return window.document.documentElement.appendChild(input); " )

  elem._execute('sendKeysToElement', {'value': [ path ], 'text': path})

  result = driver.execute_async_script( \
    "var input = arguments[0], callback = arguments[1]; "
    "var reader = new FileReader(); "
    "reader.onload = function (ev) { callback(reader.result) }; "
    "reader.onerror = function (ex) { callback(ex.message) }; "
    "reader.readAsDataURL(input.files[0]); "
    "input.remove(); "
    , elem)

  if not result.startswith('data:') :
    raise Exception("Failed to get file content: %s" % result)

  return base64.b64decode(result[result.find('base64,') + 7:])


print('Xin hãy đợi...')
if (os.path.isfile('config\\config.json') == False):
    print('Không có file cấu hình config\\config.json, chương trình sẽ kết thúc tại đây')
    time.sleep(3) #tam dung 3s roi thoat
    sys.exit("... It's over ...")

file_json = open("config\\config.json", 'r', encoding='utf-8-sig') #open file json
data_json = json.loads(file_json.read()) #returns JSON object as a dictionary
file_json.close() #close file json

#Ngày tháng năm chứng từ 
now = datetime.datetime.now()
if (data_json['ctu_tu_ngay'].strip() == '' or  data_json['ctu_den_ngay'].strip() == ''):
  ctu_tu_ngay = now.strftime('%d%m%Y')
  ctu_den_ngay = now.strftime('%d%m%Y')
else:
  ctu_tu_ngay = data_json['ctu_tu_ngay'].strip()
  ctu_den_ngay = data_json['ctu_den_ngay'].strip()

#ngay_hom_nay = now.strftime('%d%m%Y')
#ngay_hom_nay = now.strftime('21072021')
file_excel = now.strftime('%Y%m%d')

website = data_json['website']
download_folder = r"D:\APP\etax\download"
user_name = data_json['username_login_website']
password = data_json['password_login_website']
nnt_json = "{}/api/readmany/index.asp?tk=5f03cbea0a57a&cqt=22500&tin=&phong_qln=&email_qln=&phong_kk=&email_kk=&phong_kt={}&email_kt={}&phong_p=&email_p=".format(data_json['web_api'],data_json['phong_kiem_tra'],data_json['email_can_bo_kiem_tra'])
trang_thai_ctu = data_json['trang_thai_ctu']

print('Đang thực hiện xóa những file cũ...')
#delete all file from download folder
filelist = [ f for f in os.listdir(download_folder) ]
for f in filelist:
    os.remove(os.path.join(download_folder, f))

print('Lấy ds MST và cán bộ quản lý trên theone.qni.tct.vn')
with urllib.request.urlopen(nnt_json) as url:
    data = json.loads(url.read().decode())

mst = []
lst_cbql = []
for data_nnt in data['list_nnt']:
    #print(data_nnt['mst'])
    mst.append(data_nnt['mst'].strip())
    #print(data_nnt['ten_cb_kt'])
    lst_cbql.append(data_nnt['ten_cb_kt'])
#print(mst)

print('Mở trình duyệt để download file chứng từ...')
#browser = webdriver.Chrome(executable_path='chromedriver.exe')
#new option
options = webdriver.ChromeOptions()
preferences = {"download.default_directory": download_folder, "safebrowsing.enabled": "false"}
options.add_experimental_option("prefs", preferences)
browser = webdriver.Chrome(options=options)
#end new option

browser.get(website)
txt_user = browser.find_element_by_id('_userName')
txt_user.send_keys(user_name)
txt_pass = browser.find_element_by_id('password')
txt_pass.send_keys(password)
browser.find_element_by_id('dangnhap').click()

#
#vao tra cuu GNT
##1. vao menu cha
##2. vao menu con
time.sleep(2)
menu_cha = browser.find_element_by_xpath('//*[@id="menu"]/li[2]/a')
# create action chain object
action = ActionChains(browser)
# perform the operation
action.move_to_element(menu_cha).perform()
time.sleep(1)
menu_con = browser.find_element_by_xpath('//*[@id="menu"]/li[2]/ul/li[2]/a')
action.move_to_element(menu_con).click().perform()

#di chuyen con chuot ra cho khac
banner = browser.find_element_by_xpath('/html/body/div[2]')
action.move_to_element(banner).perform()

#chon iframe de lam viec
WebDriverWait(browser, 3).until(EC.frame_to_be_available_and_switch_to_it("xframe"))
#neu chon lai trang chinh:
#driver.switch_to.default_content()

chon_cqt = browser.find_element_by_xpath('//*[@id="ma_cqt"]/option[2]').click()

ngay_lap_gnt = browser.find_element_by_id('ngay_lap_tu_ngay')
ngay_lap_gnt.click() #xóa ngày lập GNT
ngay_lap_gnt.send_keys(Keys.BACK_SPACE)

browser.find_element_by_id('ngay_gui_tu_ngay').send_keys(ctu_tu_ngay)
browser.find_element_by_id('ngay_gui_den_ngay').send_keys(ctu_den_ngay)

browser.find_element_by_id('tong_tien_nt_tu').click()

#click Tra cuu: 
browser.find_element_by_xpath('//*[@id="reportForm"]/div/div/table/tbody/tr[12]/td/span/input').click()

#co bao nhieu page: 
all_page = browser.find_element_by_xpath('//*[@id="currAcc"]/b[1]').text
#print(all_page)
#co bao nhieu record: 
all_record = int(browser.find_element_by_xpath('//*[@id="currAcc"]/b[2]').text)

downloaded = 0 #ko co file download

#mst = ['0100283873','5700100256-062']
for p in range(1, int(all_page)+1):
  time.sleep(6)
  for r in range(1,51):
    try:
      stt = int(browser.find_element_by_xpath('//*[@id="allResultTableBody"]/tr['+ str(r) +']/td[1]').text)
    except Exception as e:
      stt = 0
    if (stt != 0):
      txt_mst = browser.find_element_by_xpath('//*[@id="allResultTableBody"]/tr['+ str(r) +']/td[8]').text.strip()
      #print(str(stt)+'. '+txt_mst)
      tt_ctu = browser.find_element_by_xpath('//*[@id="allResultTableBody"]/tr['+ str(r) +']/td[16]').text.strip().lower()
      if ((txt_mst in mst) and (trang_thai_ctu.strip().lower() == tt_ctu)):
        #download file
        url = browser.find_element_by_xpath('//*[@id="allResultTableBody"]/tr['+ str(r) +']/td[18]/a[2]').get_attribute('href') #get link download
        browser.execute_script('window.open("'+url+'");') #open link in new tab
        time.sleep(2)
        downloaded = 1 #da co file download
  if (p < int(all_page)):
    #go to next page
    browser.find_element_by_id('gotoPageNO_list').send_keys(str(p+1))
    browser.execute_script("gotoPage("+all_page+", 'gotoPageNO_list')")
    

browser.close() #close browser
print('Đã download xong các file chứng từ')

#Thuc hien doc file da download
if (downloaded == 1):
  print('Đang tạo file excel chứng từ theo ngày...')
  wb = openpyxl.Workbook() #tao file excel
  #tao sheet CBQL
  ws1 = wb.active
  ws1.title = 'CBQL'
  ws1['A1'] = 'MST'
  ws1['B1'] = 'CBQL'
  r = 2
  lst_len_mst = len(mst)
  for i in range(lst_len_mst):
      ws1['A'+str(r+i)] = mst[i]
      ws1['B'+str(r+i)] = lst_cbql[i]

  #ws = wb.active
  ws = wb.create_sheet(index=0,title='Chung-tu') #tao sheet
  #ws.title = "Chung-tu" #tao sheet

  ws['A1'] = 'MST'
  ws['B1'] = 'TEN NNT'
  ws['C1'] = 'TIEU MUC'
  ws['D1'] = 'SO TIEN'
  ws['E1'] = 'CB QUAN LY'
  
  filelist = [ f for f in os.listdir(download_folder) ]
  r = 2
  for f in filelist:
      detail_file = os.path.join(download_folder, f)
      #ghi du lieu ctu
      tree = ElementTree.parse(detail_file)
      #https://docs.python.org/3/library/xml.etree.elementtree.html#xml.etree.ElementTree.Element.find
      root = tree.getroot()
      mst = root.find("./NDUNG_CTU_NH/NDUNG_CTU/CHUNGTU_HDR/MST_NNOP").text #https://programmer.group/python3-standard-library-xml.etree.elementtree-xml-manipulation-api.html
      #print(mst)
      ten_nnt = root.find("./NDUNG_CTU_NH/NDUNG_CTU/CHUNGTU_HDR/TEN_NNOP").text
      #print(ten_nnt)

      for ctiet in root.findall('./NDUNG_CTU_NH/NDUNG_CTU/CHUNGTU_CTIET/ROW_CTIET'):
          tmuc = ctiet.find('MA_NDKT').text
          tien = ctiet.find('TIEN_PNOP').text
          ws['A'+str(r)] = mst
          ws['B'+str(r)] = ten_nnt
          ws['C'+str(r)] = tmuc
          ws['D'+str(r)] = float(tien)
          ws['E'+str(r)] = "=VLOOKUP(A{},'CBQL'!$A$2:$B${},2,0)".format(r,lst_len_mst+1)
          r = r+1
wb.save(filename = 'ket-qua\\ctu'+file_excel+'.xlsx') #ghi file excel
print('Đã ghi thành công file excel chứng từ theo ngày')
print('****** Kết thúc ******')
os.startfile('ket-qua\\ctu'+file_excel+'.xlsx')